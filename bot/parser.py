import json
import os
import requests
from urllib.parse import urlencode
from zipfile import ZipFile
import pandas as pd
from openpyxl import load_workbook, worksheet
from openpyxl.cell.cell import MergedCell
import gdown
from config import GOOGLE_DRIVE_URL, API_URL, YANDEX_DRIVE_URL, DataFiles
import logging


def download_schedule_file(share_link: str = GOOGLE_DRIVE_URL, output_path: str = DataFiles.SCHEDULE_FILE.value):
    """Загружает файл с расписанием с Google Docs

    Args:
        share_link (str, optional): Ссылка на Google Docs. Defaults to GOOGLE_DRIVE_URL.
        output_path (str, optional): Путь к .xlsx файлу с расписанием. Defaults to DataFiles.SCHEDULE_FILE.value.
    """
    link_split = share_link.split('/')[:6]
    link_until_id = '/'.join(link_split)

    # Формируем прямую ссылку для скачивания
    download_url = f"{link_until_id}/export?format=xlsx"
    try:
        # Скачиваем файл
        gdown.download(url=download_url, output=output_path, quiet=False, fuzzy=True)
    except Exception as e:
        logging.info(f"Ошибка при загрузке расписания: {e}")
        return
    logging.info("Успешная загрузка расписания")


def download_students_marks_from_yandex_disk(url: str = YANDEX_DRIVE_URL) -> str:
    """Загружает оценки студентов с Яндекс Диска и сохраняет в файле .json

    Args:
        url (str, optional): Ссылка на Яндекс Диск. Defaults to YANDEX_DRIVE_URL.

    Returns:
        str: Статус
    """
    try:
        final_url = API_URL + urlencode(dict(public_key=url))
        response = requests.get(final_url)
        download_url = response.json()['href']

        download_response = requests.get(download_url)
        with open('downloaded_file.zip', 'wb') as f:
            f.write(download_response.content)
        with ZipFile('downloaded_file.zip', 'r') as f:
            f.extractall('../data/xlsx')
        os.remove("downloaded_file.zip")
        parse_data_into_json()
    except Exception as e:
        logging.info(f'Ошибка при загрузке оценок: {e}')
        return str(e)
    logging.info("Успешная загрузка оценок")


def parse_data_into_json(filename: str = DataFiles.STUDENTS_FILE.value):
    """
    Парсинг данных из папки xlsx в students.json

    :return:
    """
    data_xlsx_dir = '../data/xlsx'
    students_data_as_list = []
    students_data_as_dict = {}
    dir_list = os.listdir(data_xlsx_dir)

    try:
        for dir in dir_list:
            files_list = os.listdir(data_xlsx_dir + '/' + dir)
            for file in files_list:
                excel_reader = pd.ExcelFile(data_xlsx_dir + '/' + dir + '/' + file)
                sheet_to_df_map = {}
                for sheet_name in excel_reader.sheet_names:
                    sheet_to_df_map[sheet_name] = excel_reader.parse(sheet_name, index_col=0, skiprows=2)
                    students_data_as_list += (sheet_to_df_map[sheet_name].to_dict(orient="records"))
        # Удаляем дубликаты, оставляя первое вхождение
        students_data_as_list = remove_duplicates(students_data_as_list)
        for student in students_data_as_list:
            try:
                student_id = int(student.pop('Студенч. номер'))
            except:
                continue
            students_data_as_dict[student_id] = student
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(students_data_as_dict, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f'Ошибка: {e}')


def remove_duplicates(students):
    seen = set()
    unique_students = []

    for student in students:
        student_id = student["Студенч. номер"]
        if student_id not in seen:
            seen.add(student_id)
            unique_students.append(student)

    return unique_students


def increase_column_index(index: str, increase_value: int) -> str:
    try:
        letters = ''.join(filter(str.isalpha, index))
        numbers = ''.join(filter(str.isdigit, index))
        numbers = int(numbers) + increase_value
        return letters + str(numbers)
    except Exception as e:
        print(f"Ошибка при изменении индекса:{e}")
        return


def get_cell_value(ws: worksheet, cell_coordinate: str):
    try:
        cell = ws[cell_coordinate]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell_coordinate in merged_range:
                    top_left_coordinate = merged_range.start_cell.coordinate
                    return ws[top_left_coordinate].value
        else:
            return ws[cell_coordinate].value
    except Exception as e:
        print(f"Error getting value for cell {cell_coordinate}: {e}")
        return None


def parse_schedule_column(ws: worksheet, start_index: str):
    index = start_index
    days_dict = {0: 'понедельник', 1: 'вторник', 2: 'среда', 3: 'четверг', 4: 'пятница', 5: 'суббота'}
    week_subjects = []
    for i in range(6):
        day_subjects = parse_schedule_day(ws, index)
        week_subjects.append(
            {
                'day': days_dict.get(i),
                'subjects': day_subjects
            }
        )
        index = increase_column_index(index, 11)
    return week_subjects


def parse_schedule_day(ws: worksheet, start_index: str):
    current_class = 1
    index1 = start_index
    index2 = increase_column_index(start_index, 1)
    day_subjects = []
    for i in range(5):
        cell1_value = get_cell_value(ws, index1)
        cell2_value = get_cell_value(ws, index2)
        if cell1_value == cell2_value:
            subject = '-' if cell1_value is None else cell1_value
            day_subjects.append(
                {
                    'class': current_class,
                    'subject': subject,
                    'numerator': False,
                    'denominator': False,
                    'common': True
                }
            )
        else:
            subject1 = '-' if cell1_value is None else cell1_value
            subject2 = '-' if cell2_value is None else cell2_value
            day_subjects.append(
                {
                    'class': current_class,
                    'subject': subject1,
                    'numerator': True,
                    'denominator': False,
                    'common': False
                }
            )
            day_subjects.append(
                {
                    'class': current_class,
                    'subject': subject2,
                    'numerator': False,
                    'denominator': True,
                    'common': False
                }
            )

        index1 = increase_column_index(index1, 2)
        index2 = increase_column_index(index2, 2)
        current_class += 1

    return day_subjects


def parse_schedule_by_groups(worksheet_name="Бакалавры и магистры", output_filename: str = DataFiles.GROUPS.value,
                             input_filename: str = DataFiles.SCHEDULE_FILE.value):
    wb = load_workbook(input_filename)
    ws = wb[worksheet_name]
    schedule_by_groups = []
    for row in ws.iter_rows(max_row=1, min_col=7, max_col=37):
        for cell in row:
            index = cell.coordinate
            group_name = ws[index].value
            if group_name is None:
                continue
            index = increase_column_index(index, 2)
            week_subjects = parse_schedule_column(ws, index)
            schedule_by_groups.append(
                {
                    'group_name': group_name.replace('\n', '').split('О')[0] + 'О',
                    'subjects': week_subjects
                }
            )
    with open(output_filename, 'w', encoding='utf-8') as file:
        json.dump(schedule_by_groups, file, ensure_ascii=False, indent=4)
    logging.info("Парсинг расписания по группам завершен")


def parse_classrooms(worksheet_name='аудитории', output_filename: str = DataFiles.CLASSROOMS.value,
                     input_filename: str = DataFiles.SCHEDULE_FILE.value):
    wb = load_workbook(input_filename)
    ws = wb[worksheet_name]
    classrooms = []
    for row in ws.iter_rows(max_row=1, min_col=7, max_col=34):
        for cell in row:
            index = cell.coordinate
            classroom_name = int(ws[index].value)
            if classroom_name is None:
                continue
            index = increase_column_index(index, 1)
            classroom_description = '-' if ws[index].value is None else ws[index].value
            index = increase_column_index(index, 1)

            classroom_week = parse_schedule_column(ws, index)
            classrooms.append(
                {
                    'classroom': str(classroom_name),
                    'description': classroom_description,
                    'subjects': classroom_week
                }
            )
    with open(output_filename, 'w', encoding='utf-8') as file:
        json.dump(classrooms, file, ensure_ascii=False, indent=4)
    logging.info("Парсинг расписания по аудиториям завершен")


def parse_professors(worksheet_name='Преподаватели', output_filename: str = DataFiles.PROFESSORS.value,
                     input_filename: str = DataFiles.SCHEDULE_FILE.value):
    wb = load_workbook(input_filename)
    ws = wb[worksheet_name]
    professors_list = []
    for row in ws.iter_rows(min_row=2, max_row=2, min_col=7, max_col=69):
        for cell in row:
            index = cell.coordinate
            professor_name = ws[index].value
            if professor_name is None:
                continue
            index = increase_column_index(index, 1)
            professor_schedule_week = parse_schedule_column(ws, index)
            professors_list.append(
                {
                    'professor': professor_name,
                    'subjects': professor_schedule_week
                }
            )
    with open(output_filename, 'w', encoding='utf-8') as file:
        json.dump(professors_list, file, ensure_ascii=False, indent=4)
    logging.info("Парсинг расписания по преподавателям завершен")


def parse_all_info(google_link: str = GOOGLE_DRIVE_URL, yandex_link: str = YANDEX_DRIVE_URL):
    download_schedule_file(google_link)
    download_students_marks_from_yandex_disk(yandex_link)
    parse_schedule_by_groups()
    parse_classrooms()
    parse_professors()
    schedule_file = DataFiles.SCHEDULE_FILE.value
    if os.path.exists(schedule_file):
        os.remove(schedule_file)
